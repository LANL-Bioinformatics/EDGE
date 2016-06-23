import openpyxl
import csv
from datetime import datetime
import os.path
import sys

def find_file():
    #TODO make dynamic for path and file
    #Get the excel file
    file = sys.argv[1]


    if(os.path.isfile(file)):

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        return ws
    else:
        print "Path to file incorrect. Check path"

#make active worksheet global so all methods have access without having to redo file IO operation
ws = find_file()


def get_headers():
    #need to grab the col headers and their index position so we can
    #determine the order of their columns

    cell_range = ws['A1': 'K1']  #range of header row we want to check set to 10 columns

    #need some kind of fix reference to determine order. Using what is currently in Edge
    headers = ["project","q1","q2", "Description"]

    good_cols = []
#   iterate through first row and grab column names and position
    for row in cell_range:
        for count,  cell in enumerate(row):
            temp = [str(cell.value), count+1]

            #check against col name references/headers to see if the col is one we want.
            #not algo efficent but since references is so small its no problem
            for column_name in headers:
                if(str(temp[0]).lower().strip() == column_name.lower() ):
                    good_cols.append(temp)
                else:
                    pass

    return good_cols


def get_col_data(good_cols):

#assign column letters using index position from headers
#These are integers representing col positions in the sheet
    name = None
    q1= None
    q2=None
    description=None

    for columns in good_cols:
        if(str(columns[0]).lower().strip() == "project" ):
            name = columns[1]
        elif(str(columns[0]).lower().strip() == "q1"):
            q1 = columns[1]
        elif(str(columns[0]).lower().strip() == "q2"):
            q2 = columns[1]
        elif(str(columns[0]).lower().strip() == "description"):
            description = columns[1]
        else:
            pass


#iterate through only the data we care about and put into csv format for pipeline
    good_data = []
    for rowNum, row in enumerate(ws):
        #access the excel sheet and grab the data.
            dataColName = ws.cell(row =rowNum+2, column = name).value
            dataColQ1 = ws.cell(row=rowNum + 2, column= q1).value
            dataColQ2 = ws.cell(row=rowNum + 2, column=q2).value
            dataColDes = ws.cell(row = rowNum+2, column= description).value

            temp = [dataColName, dataColQ1, dataColQ2, dataColDes]
            good_data.append(temp)

            print str(dataColName) + "," + str(dataColQ1) + "," + str(dataColQ2) + "," + str(dataColDes)

    return good_data


def output(good_data):
    #write out data into csv format
    #TODO determine file output path
    output_file = open("output.csv", 'wb')

    writer = csv.writer(output_file, dialect='excel')
    writer.writerow(list)


def main():

    starttime = datetime.now()
    get_col_data(get_headers())
    #print datetime.now() - starttime

if __name__ == '__main__':
    main()




